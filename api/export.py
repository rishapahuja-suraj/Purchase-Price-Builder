from http.server import BaseHTTPRequestHandler
import json, io
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

# ── Style constants ──────────────────────────────────────────────────────────
NAVY='091554'; BLUE='1434A4'; BLUE_L='E8EAF6'; GREY_L='F5F4F0'
WHITE='FFFFFF'; DARK='0D0D1A'; BC='C8C5BE'; RED='DC2626'; GREEN='008000'
NUM='#,##0;(#,##0);"-"'; PCT='0.0%'; MUL='0.00"x"'

def _bdr():
    t=Side(style='thin',color=BC)
    return Border(left=t,right=t,top=t,bottom=t)

def sc(ws,r,c,v=None,bold=False,sz=9,col=DARK,bg=None,ha='left',fmt=None,it=False,wrap=False):
    cell=ws.cell(row=r,column=c)
    if v is not None: cell.value=v
    cell.font=Font(name='Calibri',bold=bold,size=sz,color=col,italic=it)
    if bg: cell.fill=PatternFill('solid',fgColor=bg)
    cell.alignment=Alignment(horizontal=ha,vertical='center',wrap_text=wrap)
    if fmt: cell.number_format=fmt
    cell.border=_bdr()
    return cell

def fc(ws,r,c,formula,bold=False,col='000000',fmt=NUM,ha='right',bg=None,it=False):
    cell=ws.cell(row=r,column=c)
    cell.value=formula
    cell.font=Font(name='Calibri',bold=bold,size=9,color=col,italic=it)
    if bg: cell.fill=PatternFill('solid',fgColor=bg)
    cell.alignment=Alignment(horizontal=ha,vertical='center')
    cell.number_format=fmt
    cell.border=_bdr()
    return cell

def title_bar(ws,text,mc,row=1):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=mc)
    c=ws.cell(row,1,f'  {text}')
    c.font=Font(name='Calibri',bold=True,size=13,color=WHITE)
    c.fill=PatternFill('solid',fgColor=NAVY)
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[row].height=26

def sec_hdr(ws,r,text,mc,h=14):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=mc)
    c=ws.cell(r,1,text)
    c.font=Font(name='Calibri',bold=True,size=9,color=WHITE)
    c.fill=PatternFill('solid',fgColor=BLUE)
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[r].height=h

PERIODS=[('Stub','DATE(2026,4,1)','DATE(2026,7,1)',91),
         ('Q1',  'DATE(2026,7,1)','DATE(2026,10,1)',92),
         ('Q2',  'DATE(2026,10,1)','DATE(2027,1,1)',92),
         ('Q3',  'DATE(2027,1,1)','DATE(2027,4,1)',90),
         ('Q4',  'DATE(2027,4,1)','DATE(2027,7,1)',91)]
PYEARS=['Y2','Y3','Y4','Y5','Y6','Y7','Y8','Y9','Y10']

def _safe_int(v):
    if v is None or v == '' or isinstance(v,str): return 0
    try: return int(float(v))
    except: return 0

def _safe_float(v):
    if v is None or v == '' or isinstance(v,str): return 0.0
    try: return float(v)
    except: return 0.0

def _to_date(v):
    """Convert string/number to Python date for proper Excel date handling.
    Returns datetime if parseable, original value otherwise."""
    if v is None or v == '' or v == 'NA': return v
    if isinstance(v, (datetime, date)): return v
    if isinstance(v, str):
        s = v.strip()[:10]
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
            try:
                return datetime.strptime(s, fmt)
            except: pass
    return v


def build_excel(data):
    company   = data.get('company','Deal')
    meta      = data.get('meta','')
    is_calc   = data.get('calculatedIS',{})
    uploads   = data.get('uploads',{})
    deal      = data.get('deal',{})
    restruct  = data.get('restructuring',{})
    rev_lines = data.get('revLines',[{'k':'recRev','label':'Recurring Revenue'}])
    nr_ret    = data.get('nrRet',{})

    # Uploads as list-of-dicts
    rv_rows   = uploads.get('revenue',[]) or []
    hc_rows   = uploads.get('headcount',[]) or []
    nhc_rows  = uploads.get('nhc',[]) or []
    cf_rows   = uploads.get('cf',[]) or []
    imp_rows  = uploads.get('importCosts',[]) or []
    rest_rows = uploads.get('restructuring',[]) or []

    PCOLS=['Stub','Q1','Q2','Q3','Q4','Total Y1','Y2','Y3','Y4','Y5','Y6','Y7','Y8','Y9','Y10']
    NP=len(PCOLS); DC=2

    # IS periods → keys
    period_keys=['stub','q1','q2','q3','q4']
    year_keys  =['y1','y2','y3','y4','y5','y6','y7','y8','y9','y10']

    def is_val(period_key, line_key):
        p=is_calc.get(period_key,{}) or {}
        v=p.get(line_key,0) or 0
        if isinstance(v,str): return 0
        return int(round(float(v)/1000))

    wb=Workbook()

    # ── REVENUE DATA ────────────────────────────────────────────────────────
    ws_r=wb.active; ws_r.title='Revenue Data'
    ws_r.sheet_view.showGridLines=False

    RV_RAW_COLS=['Revenue Line','Customer Name','Contract ID','ARR ($)','Contract End',
                 'Renewal Date','Renewal Qtr','Will Renew','Up/Down Sell %','Notes']
    RV_CALC_COLS=['New ARR ($)','Stub ($)','Q1 ($)','Q2 ($)','Q3 ($)','Q4 ($)','Total Y1 ($)']
    MC_R=len(RV_RAW_COLS)+len(RV_CALC_COLS)

    for ci,w in enumerate([20,26,13,16,16,16,11,12,13,18,13]+[12]*6,1):
        ws_r.column_dimensions[gcl(ci)].width=w

    title_bar(ws_r,f'{company}  —  Revenue  |  ARR contribution per period  |  Blue=input  Black=formula',MC_R)
    ws_r.row_dimensions[2].height=10
    ws_r.merge_cells(f'A2:{gcl(MC_R)}2')
    c=ws_r['A2']; c.value='  Stub: Apr 1–Jun 30 2026 (91d)   Q1: Jul 1–Sep 30 2026 (92d)   Q2: Oct 1–Dec 31 2026 (92d)   Q3: Jan 1–Mar 31 2027 (90d)   Q4: Apr 1–Jun 30 2027 (91d)'
    c.font=Font(name='Calibri',size=7,color='888888',italic=True); c.fill=PatternFill('solid',fgColor='F2F1EE')
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws_r.row_dimensions[3].height=4

    ws_r.row_dimensions[4].height=28
    for i,h in enumerate(RV_RAW_COLS):  sc(ws_r,4,i+1,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    for i,h in enumerate(RV_CALC_COLS): sc(ws_r,4,len(RV_RAW_COLS)+i+1,h,bold=True,sz=8,col=WHITE,bg='1A3FBF',wrap=True)

    RV_DS=5; RV_N=len(rv_rows)
    kcol=len(RV_RAW_COLS)+1  # col K = New ARR
    ARR_C=gcl(4); RD_C=gcl(6); WR_C=gcl(8); UDS_C=gcl(9); K_C=gcl(kcol)

    for ri,row_d in enumerate(rv_rows):
        r=RV_DS+ri; ws_r.row_dimensions[r].height=13
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        raw_vals=[row_d.get(k,'') for k in ['Revenue Line','Customer Name','Contract ID',
            'Annualized Revenue ($)','Contract End Date (YYYY-MM-DD)','Renewal Date (YYYY-MM-DD)',
            "Renewal Quarter (e.g. Q1'28)",'Will Renew (Yes / No)',
            'Up/Down Sell % (e.g. 10 or -50)','Notes']]
        for ci2,v in enumerate(raw_vals):
            ha='right' if ci2 in [3,8] else 'left'
            # Convert date strings to datetime objects so ISNUMBER() works in Excel formulas
            if ci2 in [4,5]: v=_to_date(v)
            sc(ws_r,r,ci2+1,v,col='0000FF',sz=8,bg=alt,ha=ha)
            # Apply date format to date cells
            if ci2 in [4,5] and isinstance(ws_r.cell(row=r,column=ci2+1).value, datetime):
                ws_r.cell(row=r,column=ci2+1).number_format='YYYY-MM-DD'
        # New ARR formula
        fc(ws_r,r,kcol,f'=IF({WR_C}{r}="Yes",{ARR_C}{r}*(1+IF(AND(ABS({UDS_C}{r})<=2,{UDS_C}{r}<>0),{UDS_C}{r},{UDS_C}{r}/100)),0)',col='000000',fmt=NUM,bg=alt)
        # Period contributions
        for pi,(pname,ps,pe,pd) in enumerate(PERIODS):
            pc=kcol+1+pi
            if pname=='Stub':
                f=(f'=IF(NOT(ISNUMBER({RD_C}{r})),{ARR_C}{r}/365*{pd},'
                   f'IF({RD_C}{r}<{ps},{K_C}{r}/365*{pd},'
                   f'IF({RD_C}{r}>={pe},{ARR_C}{r}/365*{pd},'
                   f'{ARR_C}{r}/365*MAX(0,{RD_C}{r}-{ps})+{K_C}{r}/365*({pe}-{RD_C}{r}))))')
            else:
                f=(f'=IF(NOT(ISNUMBER({RD_C}{r})),{ARR_C}{r}/4,'
                   f'IF({RD_C}{r}<{ps},{K_C}{r}/4,'
                   f'IF({RD_C}{r}>={pe},{ARR_C}{r}/4,'
                   f'{ARR_C}{r}/365*MAX(0,{RD_C}{r}-{ps})+{K_C}{r}/365*({pe}-{RD_C}{r}))))')
            fc(ws_r,r,pc,f,col='000000',fmt=NUM,bg=alt)
        # Total Y1
        fc(ws_r,r,kcol+6,f'=SUM({gcl(kcol+1)}{r}:{gcl(kcol+5)}{r})',bold=True,col='000000',fmt=NUM,bg=alt)

    # Summary rows
    RV_SS=RV_DS+RV_N+2
    sec_hdr(ws_r,RV_SS,'  PERIOD SUMMARY  —  feeds Income Statement  (values in $000s)',MC_R)
    RV_SS+=1
    rv_lines_map=[
        ('Recurring Revenue',          'Recurring Revenue'),
        ('Non-Recurring Software Revenue','Non-Recurring Software Revenue'),
        ('Non-Recurring Services Revenue','Non-Recurring Services Revenue'),
    ]
    rv_sum_rows={}
    for rli,(disp,src_lbl) in enumerate(rv_lines_map):
        rs=RV_SS+rli; ws_r.row_dimensions[rs].height=15
        sc(ws_r,rs,1,disp,bold=True,sz=9,col=WHITE,bg=NAVY)
        for pi in range(5):
            pc=kcol+1+pi
            dr=f'{gcl(pc)}{RV_DS}:{gcl(pc)}{RV_DS+RV_N-1}'
            fc(ws_r,rs,pc,f'=SUMIF(A${RV_DS}:A${RV_DS+RV_N-1},"{src_lbl}",{dr})/1000',bold=True,bg='D0DCFF' if pi==4 else BLUE_L)
        fc(ws_r,rs,kcol+6,f'=SUM({gcl(kcol+1)}{rs}:{gcl(kcol+5)}{rs})',bold=True,bg='D0DCFF')
        rv_sum_rows[src_lbl]=rs

    # ── HEADCOUNT DATA ───────────────────────────────────────────────────────
    ws_h=wb.create_sheet('Headcount Data')
    ws_h.sheet_view.showGridLines=False
    HC_RAW=['Financial Line','Department','HC Category','Employment Type','Employee ID',
            'Location','Annual Comp ($)','Gross Up %','Decision','Notice Days','Exit Date']
    HC_CALC=['Gross Comp ($)','Stub ($000s)','Q1 ($000s)','Q2 ($000s)','Q3 ($000s)','Q4 ($000s)','Total Y1']
    MC_H=len(HC_RAW)+len(HC_CALC)
    for ci,w in enumerate([17,13,15,13,11,11,13,10,10,10,14]+[13]*7,1):
        ws_h.column_dimensions[gcl(ci)].width=w
    title_bar(ws_h,f'{company}  —  Headcount  |  Cost=AnnualComp×(1+GrossUp%)/365×active_days  |  Blue=input  Black=formula',MC_H)
    ws_h.row_dimensions[2].height=10
    ws_h.merge_cells(f'A2:{gcl(MC_H)}2')
    c=ws_h['A2']; c.value='  active_days = days in period before Exit Date.  Keep/NA employees contribute full period cost.  GrossUp% on top of base comp.'
    c.font=Font(name='Calibri',size=7,color='888888',italic=True); c.fill=PatternFill('solid',fgColor='F2F1EE')
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws_h.row_dimensions[3].height=4; ws_h.row_dimensions[4].height=28
    for i,h in enumerate(HC_RAW):  sc(ws_h,4,i+1,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    for i,h in enumerate(HC_CALC): sc(ws_h,4,len(HC_RAW)+i+1,h,bold=True,sz=8,col=WHITE,bg='1A3FBF',wrap=True)

    HC_DS=5; HC_N=len(hc_rows)
    GCC=len(HC_RAW)+1  # col L = Gross Comp
    AC_C=gcl(7); GU_C=gcl(8); ED_C=gcl(11); GC_C=gcl(GCC)

    for ri,row_d in enumerate(hc_rows):
        r=HC_DS+ri; ws_h.row_dimensions[r].height=13
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        raw_v=[row_d.get(k,'') for k in ['Financial Line','Department','HC Category','Employment Type',
            'Employee ID','Location (Country)','Annual Comp Base ($)','Gross Up % (auto-calculated)',
            'Decision','Notice Days','Exit Date (YYYY-MM-DD or NA)']]
        for ci2,v in enumerate(raw_v):
            if ci2 == 10 and str(v).strip().upper() != 'NA': v=_to_date(v)
            sc(ws_h,r,ci2+1,v,col='0000FF',sz=8,bg=alt,ha='right' if ci2 in [6,7,9] else 'left')
            if ci2 == 10 and isinstance(ws_h.cell(row=r,column=ci2+1).value, datetime):
                ws_h.cell(row=r,column=ci2+1).number_format='YYYY-MM-DD'
        fc(ws_h,r,GCC,f'={AC_C}{r}*(1+{GU_C}{r}/100)',col='000000',fmt=NUM,bg=alt)
        for pi,(pname,ps,pe,_) in enumerate(PERIODS):
            ef=f'IF(ISNUMBER({ED_C}{r}),{ED_C}{r},DATE(2099,1,1))'
            fc(ws_h,r,GCC+1+pi,f'={GC_C}{r}/365*MAX(0,MIN({pe},{ef})-{ps})/1000',col='000000',fmt=NUM,bg=alt)
        fc(ws_h,r,GCC+6,f'=SUM({gcl(GCC+1)}{r}:{gcl(GCC+5)}{r})',bold=True,col='000000',fmt=NUM,bg=alt)

    HC_SS=HC_DS+HC_N+2
    sec_hdr(ws_h,HC_SS,'  PERIOD SUMMARY  —  feeds Income Statement  (values in $000s)',MC_H)
    HC_SS+=1
    hc_sum_rows={}
    for hli,(disp,src_lbl) in enumerate([('HC COGS','Headcount - COGS'),('HC OPEX','Headcount - OPEX')]):
        rs=HC_SS+hli; ws_h.row_dimensions[rs].height=15
        sc(ws_h,rs,1,disp,bold=True,sz=9,col=WHITE,bg=NAVY)
        for pi in range(5):
            pc=GCC+1+pi; dr=f'{gcl(pc)}{HC_DS}:{gcl(pc)}{HC_DS+HC_N-1}'
            fc(ws_h,rs,pc,f'=SUMIF(A${HC_DS}:A${HC_DS+HC_N-1},"{src_lbl}",{dr})',bold=True,bg='D0DCFF' if pi==4 else BLUE_L)
        fc(ws_h,rs,GCC+6,f'=SUM({gcl(GCC+1)}{rs}:{gcl(GCC+5)}{rs})',bold=True,bg='D0DCFF')
        hc_sum_rows[src_lbl]=rs

    # ── NON-HEADCOUNT DATA ───────────────────────────────────────────────────
    ws_n=wb.create_sheet('Non-Headcount Data')
    ws_n.sheet_view.showGridLines=False
    NHC_RAW=['Financial Line','Department','Vendor Name','T12M Cost ($)','Decision','Term Date',
             'Util Stub','Util Q1','Util Q2','Util Q3','Util Q4','Notes']
    NHC_CALC=['Stub ($000s)','Q1 ($000s)','Q2 ($000s)','Q3 ($000s)','Q4 ($000s)','Total Y1']
    MC_N=len(NHC_RAW)+len(NHC_CALC)
    for ci,w in enumerate([17,13,22,13,12,14,10,10,10,10,10,17]+[12]*6,1):
        ws_n.column_dimensions[gcl(ci)].width=w
    title_bar(ws_n,f'{company}  —  Non-Headcount  |  Cost=T12M/365×active_days×Util%  |  Blue=input  Black=formula',MC_N)
    ws_n.row_dimensions[2].height=10
    ws_n.merge_cells(f'A2:{gcl(MC_N)}2')
    c=ws_n['A2']; c.value='  active_days = days in period before Term Date.  KEEP/blank-term rows: full period active.  Util% applied per period as entered.'
    c.font=Font(name='Calibri',size=7,color='888888',italic=True); c.fill=PatternFill('solid',fgColor='F2F1EE')
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws_n.row_dimensions[3].height=4; ws_n.row_dimensions[4].height=28
    for i,h in enumerate(NHC_RAW):  sc(ws_n,4,i+1,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    for i,h in enumerate(NHC_CALC): sc(ws_n,4,len(NHC_RAW)+i+1,h,bold=True,sz=8,col=WHITE,bg='1A3FBF',wrap=True)

    NHC_DS=5; NHC_N=len(nhc_rows)
    NHC_CC=len(NHC_RAW)+1
    T12_C=gcl(4); TD_C=gcl(6); UTIL_COLS=[gcl(7),gcl(8),gcl(9),gcl(10),gcl(11)]

    for ri,row_d in enumerate(nhc_rows):
        r=NHC_DS+ri; ws_n.row_dimensions[r].height=13
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        raw_v=[row_d.get(k,'') for k in ['Financial Line','Department','Vendor Name','T12M Cost ($)',
            'Decision (Reference Only)','Termination Date (Reference Only)',
            'Utilisation % Stub','Utilisation % Q1','Utilisation % Q2','Utilisation % Q3','Utilisation % Q4','Notes']]
        for ci2,v in enumerate(raw_v):
            if ci2 == 5: v=_to_date(v)
            sc(ws_n,r,ci2+1,v,col='0000FF',sz=8,bg=alt,ha='right' if ci2 in [3,6,7,8,9,10] else 'left')
            if ci2 == 5 and isinstance(ws_n.cell(row=r,column=ci2+1).value, datetime):
                ws_n.cell(row=r,column=ci2+1).number_format='YYYY-MM-DD'
        for pi,(pname,ps,pe,_) in enumerate(PERIODS):
            ef=f'IF(ISNUMBER({TD_C}{r}),{TD_C}{r},DATE(2099,1,1))'
            fc(ws_n,r,NHC_CC+pi,f'={T12_C}{r}/365*MAX(0,MIN({pe},{ef})-{ps})*{UTIL_COLS[pi]}{r}/1000',col='000000',fmt=NUM,bg=alt)
        fc(ws_n,r,NHC_CC+5,f'=SUM({gcl(NHC_CC)}{r}:{gcl(NHC_CC+4)}{r})',bold=True,col='000000',fmt=NUM,bg=alt)

    NHC_SS=NHC_DS+NHC_N+2
    sec_hdr(ws_n,NHC_SS,'  PERIOD SUMMARY  —  feeds Income Statement  (values in $000s)',MC_N)
    NHC_SS+=1
    nhc_sum_rows={}
    for nli,(disp,src_lbl) in enumerate([('NHC COGS','Non-Headcount - COGS'),('NHC OPEX','Non-Headcount - OPEX')]):
        rs=NHC_SS+nli; ws_n.row_dimensions[rs].height=15
        sc(ws_n,rs,1,disp,bold=True,sz=9,col=WHITE,bg=NAVY)
        for pi in range(5):
            pc=NHC_CC+pi; dr=f'{gcl(pc)}{NHC_DS}:{gcl(pc)}{NHC_DS+NHC_N-1}'
            fc(ws_n,rs,pc,f'=SUMIF(A${NHC_DS}:A${NHC_DS+NHC_N-1},"{src_lbl}",{dr})',bold=True,bg='D0DCFF' if pi==4 else BLUE_L)
        fc(ws_n,rs,NHC_CC+5,f'=SUM({gcl(NHC_CC)}{rs}:{gcl(NHC_CC+4)}{rs})',bold=True,bg='D0DCFF')
        nhc_sum_rows[src_lbl]=rs

    # ── CENTRAL FACTORY DATA ─────────────────────────────────────────────────
    ws_c=wb.create_sheet('Central Factory Data')
    ws_c.sheet_view.showGridLines=False
    CF_RAW=['Financial Line','Department','Service Name','Unit Cost ($)',
            'Units Stub','Units Q1','Units Q2','Units Q3','Units Q4','Notes']
    CF_CALC=['Stub ($000s)','Q1 ($000s)','Q2 ($000s)','Q3 ($000s)','Q4 ($000s)','Total Y1']
    MC_C=len(CF_RAW)+len(CF_CALC)
    for ci,w in enumerate([17,13,28,13,9,9,9,9,9,22]+[12]*6,1):
        ws_c.column_dimensions[gcl(ci)].width=w
    title_bar(ws_c,f'{company}  —  Central Factory  |  Cost=UnitCost×Units  |  Blue=input  Black=formula',MC_C)
    ws_c.row_dimensions[2].height=10
    ws_c.merge_cells(f'A2:{gcl(MC_C)}2')
    c=ws_c['A2']; c.value='  Unit Cost = standard Trilogy service rate.  Units = quantity of that service for the period.  Cost = UnitCost × Units per period.'
    c.font=Font(name='Calibri',size=7,color='888888',italic=True); c.fill=PatternFill('solid',fgColor='F2F1EE')
    c.alignment=Alignment(horizontal='left',vertical='center')
    ws_c.row_dimensions[3].height=4; ws_c.row_dimensions[4].height=28
    for i,h in enumerate(CF_RAW):  sc(ws_c,4,i+1,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    for i,h in enumerate(CF_CALC): sc(ws_c,4,len(CF_RAW)+i+1,h,bold=True,sz=8,col=WHITE,bg='1A3FBF',wrap=True)

    CF_DS=5; CF_N=len(cf_rows)
    CF_CC=len(CF_RAW)+1
    UC_C=gcl(4); UNIT_COLS=[gcl(5),gcl(6),gcl(7),gcl(8),gcl(9)]

    for ri,row_d in enumerate(cf_rows):
        r=CF_DS+ri; ws_c.row_dimensions[r].height=13
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        raw_v=[row_d.get(k,'') for k in ['Financial Line','Department','Service Name','Unit Cost ($)',
            'Units Stub','Units Q1','Units Q2','Units Q3','Units Q4','Rationale / Notes']]
        for ci2,v in enumerate(raw_v):
            sc(ws_c,r,ci2+1,v,col='0000FF',sz=8,bg=alt,ha='right' if ci2 in [3,4,5,6,7,8] else 'left')
        for pi,uc in enumerate(UNIT_COLS):
            fc(ws_c,r,CF_CC+pi,f'={UC_C}{r}*{uc}{r}/1000',col='000000',fmt=NUM,bg=alt)
        fc(ws_c,r,CF_CC+5,f'=SUM({gcl(CF_CC)}{r}:{gcl(CF_CC+4)}{r})',bold=True,col='000000',fmt=NUM,bg=alt)

    CF_SS=CF_DS+CF_N+2
    sec_hdr(ws_c,CF_SS,'  PERIOD SUMMARY  —  feeds Income Statement and Purchase Price  (values in $000s)',MC_C)
    CF_SS+=1
    cf_sum_rows={}
    for cli,(disp,src_lbl) in enumerate([('CF COGS','CF - COGS'),('CF OPEX','CF - OPEX')]):
        rs=CF_SS+cli; ws_c.row_dimensions[rs].height=15
        sc(ws_c,rs,1,disp,bold=True,sz=9,col=WHITE,bg=NAVY)
        for pi in range(5):
            pc=CF_CC+pi; dr=f'{gcl(pc)}{CF_DS}:{gcl(pc)}{CF_DS+CF_N-1}'
            fc(ws_c,rs,pc,f'=SUMIF(A${CF_DS}:A${CF_DS+CF_N-1},"{src_lbl}",{dr})',bold=True,bg='D0DCFF' if pi==4 else BLUE_L)
        fc(ws_c,rs,CF_CC+5,f'=SUM({gcl(CF_CC)}{rs}:{gcl(CF_CC+4)}{rs})',bold=True,bg='D0DCFF')
        cf_sum_rows[src_lbl]=rs

    # ── IMPORT COSTS ─────────────────────────────────────────────────────────
    ws_i=wb.create_sheet('Import Costs Data')
    ws_i.sheet_view.showGridLines=False
    for ci,w in enumerate([17,26,13,10,13,15,26],1): ws_i.column_dimensions[gcl(ci)].width=w
    title_bar(ws_i,f'{company}  —  Import Costs  |  Total Cost = Units × Unit Price',7)
    ws_i.row_dimensions[2].height=5; ws_i.row_dimensions[3].height=15
    for ci,h in enumerate(['Department','Import Service','Import Driver','Units','Unit Price ($)','Total Cost ($)','Rationale'],1):
        sc(ws_i,3,ci,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    IMP_DS=4; IMP_N=len(imp_rows)
    for ri,row_d in enumerate(imp_rows):
        r=IMP_DS+ri; ws_i.row_dimensions[r].height=13
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        for ci2,k in enumerate(['Department','Import Service','Import Driver','No. of Units','Unit Price']):
            sc(ws_i,r,ci2+1,row_d.get(k,''),col='0000FF',sz=8,bg=alt,ha='right' if ci2 in [3,4] else 'left')
        fc(ws_i,r,6,f'=D{r}*E{r}',col='000000',fmt=NUM,bg=alt)
        sc(ws_i,r,7,row_d.get('Rationale',''),col='555555',sz=8,bg=alt,it=True)
    IMP_TOT=IMP_DS+IMP_N+1
    ws_i.row_dimensions[IMP_TOT].height=15
    sc(ws_i,IMP_TOT,1,'TOTAL',bold=True,sz=9,col=WHITE,bg=NAVY)
    fc(ws_i,IMP_TOT,6,f'=SUM(F{IMP_DS}:F{IMP_DS+IMP_N-1})',bold=True,bg='D0DCFF')

    # ── RESTRUCTURING DATA ───────────────────────────────────────────────────
    ws_rest=wb.create_sheet('Restructuring Data')
    ws_rest.sheet_view.showGridLines=False
    MONTHS=[f'Month {i}' for i in range(1,13)]+['>12 Months']
    REST_HDRS=['Cost Item','Basis of','Amount ($)','Reasoning / Comments']+MONTHS+['SUM Check']
    MC_REST=len(REST_HDRS)
    for ci,w in enumerate([22,12,13,30]+[8]*13+[10],1): ws_rest.column_dimensions[gcl(ci)].width=w
    title_bar(ws_rest,f'{company}  —  Mgmt Restructuring  |  Amount must equal SUM of monthly columns',MC_REST)
    ws_rest.row_dimensions[2].height=5; ws_rest.row_dimensions[3].height=18
    for ci,h in enumerate(REST_HDRS): sc(ws_rest,3,ci+1,h,bold=True,sz=8,col=WHITE,bg=NAVY,wrap=True)
    REST_DS=4; REST_N=len(rest_rows)
    m_keys=['Month 1','Month 2','Month 3','Month 4','Month 5','Month 6','Month 7',
            'Month 8','Month 9','Month 10','Month 11','Month 12','>12 Months']
    for ri,row_d in enumerate(rest_rows):
        r=REST_DS+ri; ws_rest.row_dimensions[r].height=14
        alt=GREY_L if ri%2==0 else 'FFFFFF'
        sc(ws_rest,r,1,row_d.get('Cost Item',''),col='0000FF',sz=8,bg=alt)
        sc(ws_rest,r,2,row_d.get('Basis of',''),col='555555',sz=8,bg=alt,ha='center')
        sc(ws_rest,r,3,_safe_float(row_d.get('Amount ($)',0)),col='0000FF',sz=8,bg=alt,fmt=NUM,ha='right')
        sc(ws_rest,r,4,row_d.get('Reasoning / Comments',''),col='555555',sz=8,bg=alt,it=True)
        for mi,mk in enumerate(m_keys):
            sc(ws_rest,r,5+mi,_safe_float(row_d.get(mk,0)),col='0000FF',sz=8,bg=alt,fmt=NUM,ha='right')
        # SUM check = Amount - SUM(months) → should be 0
        month_range=f'E{r}:{gcl(5+12)}{r}'
        fc(ws_rest,r,MC_REST,f'=C{r}-SUM({month_range})',col=RED,fmt=NUM,bg=alt)
    REST_TOT=REST_DS+REST_N+1
    ws_rest.row_dimensions[REST_TOT].height=15
    sc(ws_rest,REST_TOT,1,'TOTAL RESTRUCTURING',bold=True,sz=9,col=WHITE,bg=NAVY)
    fc(ws_rest,REST_TOT,3,f'=SUM(C{REST_DS}:C{REST_DS+REST_N-1})',bold=True,fmt=NUM,bg='D0DCFF')

    # ── PURCHASE PRICE ───────────────────────────────────────────────────────
    ws_pp=wb.create_sheet('Purchase Price')
    ws_pp.sheet_view.showGridLines=False
    for ci,w in enumerate([36,16,10,26],1): ws_pp.column_dimensions[gcl(ci)].width=w
    title_bar(ws_pp,f'{company}  —  Purchase Price Waterfall',4)
    ws_pp.row_dimensions[2].height=6; pp_r=3
    sec_hdr(ws_pp,pp_r,"  PRICE ADJUSTMENTS  |  Amounts in $'000s",4,15); pp_r+=1
    for ci2,h in enumerate(['Cost Item',"Amount ($'000s)",'Tag','Notes'],1):
        sc(ws_pp,pp_r,ci2,h,bold=True,sz=9,col=WHITE,bg=BLUE,ha='right' if ci2==2 else 'center' if ci2==3 else 'left')
    pp_r+=1

    pp_fields=[
        ('Base Purchase Price (Net of Cash)', restruct.get('pp',0)/1000, 'BASE'),
        ('Working Capital Adjustment',        restruct.get('wc',0)/1000, ''),
        ('Management Restructuring Costs',    restruct.get('restruct',0)/1000, ''),
        ('Import Cost',                       None, '←  Import Costs Data'),
        ('Lease Adjustment',                  restruct.get('lease',0)/1000, ''),
        ('Employee Cost Adjustment',          restruct.get('emp',0)/1000, ''),
    ]
    pp_item_rows=[]
    for lbl2,val2,tag2 in pp_fields:
        alt=GREY_L if pp_r%2==0 else 'FFFFFF'
        sc(ws_pp,pp_r,1,lbl2,bg=alt)
        if lbl2=='Import Cost':
            fc(ws_pp,pp_r,2,f"='Import Costs Data'!F{IMP_TOT}/1000000",col=GREEN,fmt=NUM,bg=alt)
        else:
            sc(ws_pp,pp_r,2,round(float(val2),3) if val2 else 0,col='0000FF',bg=alt,fmt=NUM,ha='right')
        sc(ws_pp,pp_r,3,tag2,sz=8,col='666666',ha='center',bg=alt)
        sc(ws_pp,pp_r,4,'',bg=alt)
        pp_item_rows.append(pp_r); pp_r+=1

    sc(ws_pp,pp_r,1,'Adjusted Purchase Price',bold=True,col=WHITE,bg=NAVY)
    for ci2 in [3,4]: ws_pp.cell(pp_r,ci2).fill=PatternFill('solid',fgColor=NAVY)
    c2=ws_pp.cell(pp_r,2)
    c2.value=f'=SUM(B{pp_item_rows[0]}:B{pp_r-1})'
    c2.font=Font(name='Calibri',bold=True,color=WHITE)
    c2.fill=PatternFill('solid',fgColor=NAVY); c2.number_format=NUM
    c2.alignment=Alignment(horizontal='right',vertical='center'); c2.border=_bdr()
    adj_row=pp_r; pp_r+=2

    sec_hdr(ws_pp,pp_r,'  INVESTMENT MULTIPLES  |  Y1 Revenue cross-linked from Income Statement',4,15); pp_r+=1
    y1_rev_is_row=12  # will update after IS built
    mult_start=pp_r

    # ── INCOME STATEMENT ─────────────────────────────────────────────────────
    ws_is=wb.create_sheet('Income Statement')
    ws_is.sheet_view.showGridLines=False
    ws_is.freeze_panes='B8'
    for ci,w in enumerate([32]+[12]*5+[13]+[11]*9+[16],1):
        ws_is.column_dimensions[gcl(ci)].width=w

    title_bar(ws_is,f'{company}  —  Performance Plan Builder  |  Income Statement',NP+DC)
    ws_is.row_dimensions[2].height=12
    ws_is.merge_cells(f'A2:{gcl(NP+DC)}2')
    c=ws_is['A2']; c.value=f'  {meta}'
    c.font=Font(name='Calibri',size=8,color='AAAAAA',italic=True)
    c.fill=PatternFill('solid',fgColor='1A1A2E'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws_is.row_dimensions[3].height=10
    ws_is.merge_cells(f'A3:{gcl(NP+DC)}3')
    c=ws_is['A3']; c.value="  Values in $'000s   |   Blue=input   Black=formula   Green=cross-sheet link"
    c.font=Font(name='Calibri',size=7,color='999999',italic=True)
    c.fill=PatternFill('solid',fgColor='F2F1EE'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws_is.row_dimensions[4].height=5; ws_is.row_dimensions[5].height=15
    sc(ws_is,5,1,'Line Item',bold=True,sz=9,col=WHITE,bg=NAVY)
    for i,p in enumerate(PCOLS):
        sc(ws_is,5,DC+i,p,bold=True,sz=9,col=WHITE,bg='1A3FBF' if p=='Total Y1' else NAVY,ha='right')
    sc(ws_is,5,NP+DC,'Source',bold=True,sz=9,col=WHITE,bg=NAVY,ha='center')
    ws_is.row_dimensions[6].height=10
    sc(ws_is,6,1,'',bg=BLUE_L)
    for i in range(NP): sc(ws_is,6,DC+i,"($'000s)",it=True,sz=7,col='BBBBBB',bg=BLUE_L,ha='right')
    sc(ws_is,6,NP+DC,'',bg=BLUE_L); ws_is.row_dimensions[7].height=4

    is_row_map={}; cur_is=8

    def write_sec(lbl):
        nonlocal cur_is
        ws_is.merge_cells(start_row=cur_is,start_column=1,end_row=cur_is,end_column=NP+DC)
        c=ws_is.cell(cur_is,1,f'  {lbl}')
        c.font=Font(name='Calibri',bold=True,size=9,color=WHITE)
        c.fill=PatternFill('solid',fgColor=BLUE)
        c.alignment=Alignment(horizontal='left',vertical='center')
        ws_is.row_dimensions[cur_is].height=14
        is_row_map[lbl]=cur_is; cur_is+=1

    def write_data_row(label,stubs_formulas,src_lbl,isSub=False,isResult=False,isPct=False,y2_key=None):
        nonlocal cur_is
        alt=GREY_L if cur_is%2==0 else 'FFFFFF'
        bg_r='E8EAF6' if isSub or isResult else alt
        bold=isSub or isResult
        ws_is.row_dimensions[cur_is].height=14
        sc(ws_is,cur_is,1,('  ' if bold else '    ')+label.strip(),bold=bold,it=isPct,sz=9,
           col='555555' if isPct else DARK,bg=bg_r)
        # Stub–Q4 from formulas
        for pi,f in enumerate(stubs_formulas[:5]):
            bg2='D0DCFF' if pi==5 else bg_r
            col2=GREEN if f and f.startswith("='") else '000000'
            fc(ws_is,cur_is,DC+pi,f,bold=bold,col=col2,fmt=PCT if isPct else NUM,bg=bg_r)
        # Total Y1
        fc(ws_is,cur_is,DC+5,f'=SUM({gcl(DC)}{cur_is}:{gcl(DC+4)}{cur_is})',bold=bold,col='000000',fmt=PCT if isPct else NUM,bg='D0DCFF')
        # Y2-Y10 from IS calc
        key=label.strip()
        for i,yk in enumerate(PYEARS):
            v=0
            pk=is_calc.get(yk,{})
            if pk:
                lk=None
                if 'Revenue' in key or 'Recurring' in key:
                    lk='totRev' if 'Total' in key else ('nrSoft' if 'Software' in key else ('nrServ' if 'Service' in key else 'recRev'))
                elif 'COGS' in key and 'HC' not in key and 'Non-Head' not in key and 'CF' not in key: lk='totCogs'
                elif 'Headcount' in key and 'COGS' in key: lk='hcCogs'
                elif 'Non-Headcount' in key and 'COGS' in key: lk='nhcCogs'
                elif 'Central Factory' in key and 'COGS' in key: lk='cfCogs'
                elif 'Gross Profit' == key: lk='gp'
                elif 'Headcount' in key and 'OPEX' in key: lk='hcOpex'
                elif 'Non-Headcount' in key and 'OPEX' in key: lk='nhcOpex'
                elif 'Write' in key: lk='revWO'
                elif 'Central Factory' in key and 'OPEX' in key: lk='cfOpex'
                elif 'Core' in key: lk='coreAlloc'
                elif 'Total OPEX' == key: lk='totOpex'
                elif 'EBITDA' == key: lk='ebitda'
                elif 'Non-Recurring Software' in key: lk='nrSoft'
                elif 'Non-Recurring Services' in key: lk='nrServ'
                elif 'Total Revenue' == key: lk='totRev'
                if lk: v=_safe_int(pk.get(lk,0))
            if isPct:
                raw=_safe_float(pk.get('gpMgn',0) if 'Gross Margin' in key else pk.get('netMgn',0))
                v2=raw/100 if abs(raw)<5 else raw/100
                sc(ws_is,cur_is,DC+6+i,v2,it=True,sz=8,col='666666',bg=bg_r,fmt=PCT,ha='right')
            else:
                sc(ws_is,cur_is,DC+6+i,v,sz=9,col='0000FF',bg=bg_r,fmt=NUM,ha='right')
        sc(ws_is,cur_is,NP+DC,src_lbl,it=True,sz=7,col='999999',bg=bg_r,ha='center')
        is_row_map[label.strip()]=cur_is; cur_is+=1

    def cs(sheet,row,col_idx):
        return f"='{sheet}'!{gcl(col_idx)}{row}"

    def rv_f(src_lbl): return [cs('Revenue Data',rv_sum_rows[src_lbl],kcol+1+pi) for pi in range(5)]
    def hc_f(src_lbl): return [cs('Headcount Data',hc_sum_rows[src_lbl],GCC+1+pi) for pi in range(5)]
    def nhc_f(src_lbl): return [cs('Non-Headcount Data',nhc_sum_rows[src_lbl],NHC_CC+pi) for pi in range(5)]
    def cf_f(src_lbl):  return [cs('Central Factory Data',cf_sum_rows[src_lbl],CF_CC+pi) for pi in range(5)]
    def is_val_f(lk):   return [f'={_safe_int(is_calc.get(pk,{}).get(lk,0))}' for pk in period_keys]

    write_sec('REVENUE')
    write_data_row('Recurring Revenue',          rv_f('Recurring Revenue'),          'Revenue Data')
    write_data_row('Non-Recurring Software Revenue',rv_f('Non-Recurring Software Revenue'),'Revenue Data')
    write_data_row('Non-Recurring Services Revenue',rv_f('Non-Recurring Services Revenue'),'Revenue Data')

    # Total Revenue
    tr_r=cur_is; ws_is.row_dimensions[tr_r].height=14
    rev_comps=[is_row_map['Recurring Revenue'],is_row_map['Non-Recurring Software Revenue'],is_row_map['Non-Recurring Services Revenue']]
    sc(ws_is,tr_r,1,'  Total Revenue',bold=True,sz=9,bg='E8EAF6')
    for pi in range(5):
        f='+'.join(gcl(DC+pi)+str(r2) for r2 in rev_comps)
        fc(ws_is,tr_r,DC+pi,f'={f}',bold=True,fmt=NUM,bg='E8EAF6')
    fc(ws_is,tr_r,DC+5,f'=SUM({gcl(DC)}{tr_r}:{gcl(DC+4)}{tr_r})',bold=True,fmt=NUM,bg='D0DCFF')
    for i in range(9):
        sc(ws_is,tr_r,DC+6+i,_safe_int(is_calc.get(PYEARS[i].lower(),{}).get('totRev',0)),bold=True,col='000000',bg='E8EAF6',fmt=NUM,ha='right')
    sc(ws_is,tr_r,NP+DC,'Formula',it=True,sz=7,col='999999',bg='E8EAF6',ha='center')
    is_row_map['Total Revenue']=tr_r; cur_is+=1

    write_sec('COST OF GOODS SOLD')
    write_data_row('Headcount — COGS',       hc_f('Headcount - COGS'),     'Headcount Data')
    write_data_row('Non-Headcount — COGS',   nhc_f('Non-Headcount - COGS'),'Non-Headcount Data')
    write_data_row('Central Factory — COGS', cf_f('CF - COGS'),            'Central Factory Data')

    tc_r=cur_is; ws_is.row_dimensions[tc_r].height=14
    cogs_comps=[is_row_map.get(k) for k in ['Headcount — COGS','Non-Headcount — COGS','Central Factory — COGS'] if is_row_map.get(k)]
    sc(ws_is,tc_r,1,'  Total COGS',bold=True,sz=9,bg='E8EAF6')
    for pi in range(5):
        fc(ws_is,tc_r,DC+pi,f'={"+".join(gcl(DC+pi)+str(r2) for r2 in cogs_comps)}',bold=True,fmt=NUM,bg='E8EAF6')
    fc(ws_is,tc_r,DC+5,f'=SUM({gcl(DC)}{tc_r}:{gcl(DC+4)}{tc_r})',bold=True,fmt=NUM,bg='D0DCFF')
    for i in range(9):
        sc(ws_is,tc_r,DC+6+i,_safe_int(is_calc.get(PYEARS[i].lower(),{}).get('totCogs',0)),bold=True,col='000000',bg='E8EAF6',fmt=NUM,ha='right')
    sc(ws_is,tc_r,NP+DC,'Formula',it=True,sz=7,col='999999',bg='E8EAF6',ha='center')
    is_row_map['Total COGS']=tc_r; cur_is+=1

    # Gross Profit
    gp_r=cur_is; r_tr2=is_row_map['Total Revenue']; r_tc2=is_row_map['Total COGS']
    ws_is.row_dimensions[gp_r].height=15
    sc(ws_is,gp_r,1,'  Gross Profit',bold=True,sz=10,col=WHITE,bg=NAVY)
    for pi in range(5):
        col=gcl(DC+pi)
        fc(ws_is,gp_r,DC+pi,f'={col}{r_tr2}-{col}{r_tc2}',bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    fc(ws_is,gp_r,DC+5,f'=SUM({gcl(DC)}{gp_r}:{gcl(DC+4)}{gp_r})',bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    for i in range(9):
        fc(ws_is,gp_r,DC+6+i,_safe_int(is_calc.get(PYEARS[i].lower(),{}).get('gp',0)),bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    ws_is.cell(gp_r,NP+DC).fill=PatternFill('solid',fgColor=NAVY)
    is_row_map['Gross Profit']=gp_r; cur_is+=1

    # Gross Margin %
    gm_r=cur_is; ws_is.row_dimensions[gm_r].height=13
    sc(ws_is,gm_r,1,'  Gross Margin %',it=True,sz=8,col='666666',bg=GREY_L)
    for pi in range(5):
        col=gcl(DC+pi)
        fc(ws_is,gm_r,DC+pi,f'=IFERROR({col}{gp_r}/{col}{tr_r},0)',it=True,col='666666',fmt=PCT,bg=GREY_L)
    fc(ws_is,gm_r,DC+5,f'=IFERROR({gcl(DC+5)}{gp_r}/{gcl(DC+5)}{tr_r},0)',it=True,col='666666',fmt=PCT,bg='D0DCFF')
    for i in range(9):
        raw=_safe_float(is_calc.get(PYEARS[i].lower(),{}).get('gpMgn',0))
        sc(ws_is,gm_r,DC+6+i,raw/100 if abs(raw)<5 else raw/100,it=True,sz=8,col='666666',bg=GREY_L,fmt=PCT,ha='right')
    sc(ws_is,gm_r,NP+DC,'Formula',it=True,sz=7,col='999999',bg=GREY_L,ha='center'); cur_is+=1

    write_sec('OPERATING EXPENSES')
    write_data_row('Headcount — OPEX',       hc_f('Headcount - OPEX'),     'Headcount Data')
    write_data_row('Non-Headcount — OPEX',   nhc_f('Non-Headcount - OPEX'),'Non-Headcount Data')
    write_data_row('Revenue Write-Off (0.5%)',is_val_f('revWO'),            'Model')
    write_data_row('Central Factory — OPEX', cf_f('CF - OPEX'),            'Central Factory Data')
    write_data_row('Core Allocation (5.5%)', is_val_f('coreAlloc'),        'Model')

    to_r=cur_is; opex_comps=[is_row_map.get(k) for k in ['Headcount — OPEX','Non-Headcount — OPEX','Revenue Write-Off (0.5%)','Central Factory — OPEX','Core Allocation (5.5%)'] if is_row_map.get(k)]
    ws_is.row_dimensions[to_r].height=14; sc(ws_is,to_r,1,'  Total OPEX',bold=True,sz=9,bg='E8EAF6')
    for pi in range(5):
        fc(ws_is,to_r,DC+pi,f'={"+".join(gcl(DC+pi)+str(r2) for r2 in opex_comps)}',bold=True,fmt=NUM,bg='E8EAF6')
    fc(ws_is,to_r,DC+5,f'=SUM({gcl(DC)}{to_r}:{gcl(DC+4)}{to_r})',bold=True,fmt=NUM,bg='D0DCFF')
    for i in range(9):
        sc(ws_is,to_r,DC+6+i,_safe_int(is_calc.get(PYEARS[i].lower(),{}).get('totOpex',0)),bold=True,col='000000',bg='E8EAF6',fmt=NUM,ha='right')
    sc(ws_is,to_r,NP+DC,'Formula',it=True,sz=7,col='999999',bg='E8EAF6',ha='center')
    is_row_map['Total OPEX']=to_r; cur_is+=1

    # EBITDA
    eb_r=cur_is; r_gp2=is_row_map['Gross Profit']; r_to2=is_row_map['Total OPEX']
    ws_is.row_dimensions[eb_r].height=16
    sc(ws_is,eb_r,1,'  EBITDA',bold=True,sz=11,col=WHITE,bg=NAVY)
    for pi in range(5):
        col=gcl(DC+pi)
        fc(ws_is,eb_r,DC+pi,f'={col}{r_gp2}-{col}{r_to2}',bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    fc(ws_is,eb_r,DC+5,f'=SUM({gcl(DC)}{eb_r}:{gcl(DC+4)}{eb_r})',bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    for i in range(9):
        fc(ws_is,eb_r,DC+6+i,_safe_int(is_calc.get(PYEARS[i].lower(),{}).get('ebitda',0)),bold=True,col=WHITE,fmt=NUM,bg=NAVY)
    ws_is.cell(eb_r,NP+DC).fill=PatternFill('solid',fgColor=NAVY)
    is_row_map['EBITDA']=eb_r; cur_is+=1

    # EBITDA Margin %
    em_r=cur_is; ws_is.row_dimensions[em_r].height=13
    sc(ws_is,em_r,1,'  EBITDA Margin %',it=True,sz=8,col='666666',bg=GREY_L)
    for pi in range(5):
        col=gcl(DC+pi)
        fc(ws_is,em_r,DC+pi,f'=IFERROR({col}{eb_r}/{col}{tr_r},0)',it=True,col='666666',fmt=PCT,bg=GREY_L)
    fc(ws_is,em_r,DC+5,f'=IFERROR({gcl(DC+5)}{eb_r}/{gcl(DC+5)}{tr_r},0)',it=True,col='666666',fmt=PCT,bg='D0DCFF')
    for i in range(9):
        raw=_safe_float(is_calc.get(PYEARS[i].lower(),{}).get('netMgn',0))
        sc(ws_is,em_r,DC+6+i,raw/100 if abs(raw)<5 else raw/100,it=True,sz=8,col='666666',bg=GREY_L,fmt=PCT,ha='right')
    sc(ws_is,em_r,NP+DC,'Formula',it=True,sz=7,col='999999',bg=GREY_L,ha='center'); cur_is+=1

    # ── Complete Purchase Price multiples with IS reference ──────────────────
    y1_rev_is_row=is_row_map.get('Total Revenue',12)
    for lbl3,formula3,fmt3,clr3 in [
        ('Adjusted Purchase Price',     f'=B{adj_row}',                                NUM,'000000'),
        ("Y1 Revenue  ←  Income Statement",f"='Income Statement'!G{y1_rev_is_row}",   NUM, GREEN),
        ('EV / Y1 Revenue',             None,                                           MUL,'000000')]:
        sc(ws_pp,mult_start,1,lbl3,bold=lbl3.startswith('EV'),bg=GREY_L)
        f3=f'=IFERROR(B{mult_start-2}/B{mult_start-1},0)' if lbl3.startswith('EV') else formula3
        c2=ws_pp.cell(mult_start,2); c2.value=f3
        c2.font=Font(name='Calibri',bold=lbl3.startswith('EV'),color=clr3,size=9)
        c2.number_format=fmt3; c2.alignment=Alignment(horizontal='right',vertical='center')
        c2.fill=PatternFill('solid',fgColor=GREY_L); c2.border=_bdr()
        mult_start+=1

    # ── IRR-DCF ──────────────────────────────────────────────────────────────
    ws2=wb.create_sheet('IRR-DCF')
    ws2.sheet_view.showGridLines=False
    for ci,w in enumerate([22,16,16,14,18,18],1): ws2.column_dimensions[gcl(ci)].width=w
    pp_val_k=abs(_safe_float(restruct.get('adjPP',0)))/1000 or abs(_safe_float(deal.get('purchasePrice',0)))/1000
    title_bar(ws2,f'{company}  —  IRR / DCF  |  Y1 Revenue + EBITDA cross-linked from Income Statement',6)
    ws2.row_dimensions[2].height=12; ws2.merge_cells('A2:F2')
    c=ws2['A2']; c.value=f'  Tax Rate: {deal.get("taxRate",5)}%  |  PP: {company}  |  Revenue and EBITDA for Y1 pull live from Income Statement tab'
    c.font=Font(name='Calibri',size=8,color='AAAAAA',italic=True)
    c.fill=PatternFill('solid',fgColor='1A1A2E'); c.alignment=Alignment(horizontal='left',vertical='center')
    ws2.row_dimensions[3].height=6; ws2.row_dimensions[4].height=15
    for i,h in enumerate(['Period',"Revenue ($'000s)","EBITDA ($'000s)","Tax ($'000s)","After-Tax CF ($'000s)","Cumul. CF ($'000s)"],1):
        sc(ws2,4,i,h,bold=True,sz=9,col=WHITE,bg=NAVY,ha='center' if i>1 else 'left')
    ws2.row_dimensions[5].height=14
    sc(ws2,5,1,'Day 0  (Investment)',bold=True,bg=GREY_L)
    for ci2 in [2,3,4]: sc(ws2,5,ci2,'—',ha='center',bg=GREY_L)
    sc(ws2,5,5,-pp_val_k,bold=True,col=RED,fmt=NUM,ha='right',bg=GREY_L)
    sc(ws2,5,6,-pp_val_k,bold=True,col=RED,fmt=NUM,ha='right',bg=GREY_L)
    tr_is=is_row_map.get('Total Revenue',12); eb_is=is_row_map.get('EBITDA',28)
    tax_rate=_safe_float(deal.get('taxRate',5))/100
    for yi in range(10):
        er=6+yi; alt=GREY_L if yi%2==0 else 'FFFFFF'; ws2.row_dimensions[er].height=14
        sc(ws2,er,1,f'Year {yi+1}',bg=alt)
        if yi==0:
            c2=ws2.cell(er,2); c2.value=f"='Income Statement'!G{tr_is}"
            c2.font=Font(name='Calibri',color=GREEN,size=9); c2.number_format=NUM
            c2.alignment=Alignment(horizontal='right',vertical='center'); c2.fill=PatternFill('solid',fgColor=alt); c2.border=_bdr()
            c3=ws2.cell(er,3); c3.value=f"='Income Statement'!G{eb_is}"
            c3.font=Font(name='Calibri',color=GREEN,size=9); c3.number_format=NUM
            c3.alignment=Alignment(horizontal='right',vertical='center'); c3.fill=PatternFill('solid',fgColor=alt); c3.border=_bdr()
        else:
            yk=PYEARS[yi-1].lower()
            sc(ws2,er,2,_safe_int(is_calc.get(yk,{}).get('totRev',0)),col='0000FF',fmt=NUM,ha='right',bg=alt)
            sc(ws2,er,3,_safe_int(is_calc.get(yk,{}).get('ebitda',0)),col='0000FF',fmt=NUM,ha='right',bg=alt)
        for ci2,f2,bold2 in [(4,f'=ROUND(MAX(0,C{er})*{tax_rate},0)',False),(5,f'=MAX(0,C{er})-D{er}',False),(6,f'=F{er-1}+E{er}',True)]:
            c2=ws2.cell(er,ci2); c2.value=f2
            c2.font=Font(name='Calibri',bold=bold2,color='000000',size=9); c2.number_format=NUM
            c2.alignment=Alignment(horizontal='right',vertical='center'); c2.fill=PatternFill('solid',fgColor=alt); c2.border=_bdr()
    last_yr=15; ws2.row_dimensions[last_yr+1].height=8; kpi_r=last_yr+2
    sec_hdr(ws2,kpi_r,'  KEY METRICS',6,15)
    for ki,(lbl,f2,fmt2,bold2) in enumerate([
        ('IRR  (10-Year)',f'=IRR(F5:F{last_yr})','0.0%',True),
        ('MoM  (10-Year  gross CF / investment)',f'=IFERROR(SUM(E6:E{last_yr})/ABS(E5),0)',MUL,True),
        ('Investment (Day 0)',f'=ABS(E5)',NUM,False)]):
        r2=kpi_r+1+ki; ws2.row_dimensions[r2].height=14
        sc(ws2,r2,1,lbl,bold=bold2,bg=GREY_L)
        c2=ws2.cell(r2,2); c2.value=f2
        c2.font=Font(name='Calibri',bold=bold2,color='000000',size=9); c2.number_format=fmt2
        c2.alignment=Alignment(horizontal='right',vertical='center'); c2.fill=PatternFill('solid',fgColor=GREY_L); c2.border=_bdr()

    # ── COVER ────────────────────────────────────────────────────────────────
    cov=wb.create_sheet('Cover',0); cov.sheet_view.showGridLines=False
    for col,w in [('A',4),('B',38),('C',36)]: cov.column_dimensions[col].width=w
    cov.row_dimensions[2].height=48; cov.row_dimensions[3].height=20; cov.row_dimensions[4].height=8
    cov.merge_cells('B2:C2'); c=cov['B2']; c.value=company
    c.font=Font(name='Calibri',bold=True,size=30,color=NAVY); c.alignment=Alignment(horizontal='left',vertical='center')
    cov.merge_cells('B3:C3'); c=cov['B3']; c.value='Performance Plan  |  Deal Financial Model'
    c.font=Font(name='Calibri',size=13,color='444444'); c.alignment=Alignment(horizontal='left',vertical='center')
    cov.row_dimensions[5].height=14; cov.merge_cells('B5:C5'); c=cov['B5']
    c.value=f'  {meta}'
    c.font=Font(name='Calibri',size=9,color='777777',italic=True); c.alignment=Alignment(horizontal='left',vertical='center')
    cov.row_dimensions[6].height=10; cov.row_dimensions[7].height=15
    cov.merge_cells('B7:C7'); c=cov['B7']; c.value='SHEET CONTENTS'
    c.font=Font(name='Calibri',bold=True,size=10,color=NAVY); c.alignment=Alignment(horizontal='left',vertical='center')
    tabs_cov=[
        ('Income Statement','Full P&L  Stub→Q4→Y1–Y10  |  Cross-sheet formulas from Revenue/HC/NHC/CF tabs'),
        ('Revenue Data',    f'ARR per customer per period  |  New ARR formula  |  SUMIF summary  ({len(rv_rows)} rows)'),
        ('Headcount Data',  f'AnnualComp×GrossUp%/365×active_days  |  SUMIF summary  ({len(hc_rows)} rows)'),
        ('Non-Headcount Data',f'T12M/365×active_days×Util%  |  SUMIF summary  ({len(nhc_rows)} rows)'),
        ('Central Factory Data',f'UnitCost×Units  |  SUMIF summary  ({len(cf_rows)} rows)'),
        ('Import Costs Data',f'Units×UnitPrice formula  |  Total feeds Purchase Price  ({len(imp_rows)} rows)'),
        ('Restructuring Data',f'Monthly cashflow schedule  |  SUM check column  ({len(rest_rows)} rows)'),
        ('Purchase Price',  'Adjusted PP waterfall  |  Cross-sheet link to Import Costs and IS'),
        ('IRR-DCF',         'Y1 Revenue+EBITDA from IS  |  Excel IRR() formula  |  MoM multiple'),
    ]
    for i,(tab,desc) in enumerate(tabs_cov):
        r2=8+i; cov.row_dimensions[r2].height=14
        cov.cell(r2,2,f'  →  {tab}').font=Font(name='Calibri',bold=True,size=9,color=BLUE)
        cov.cell(r2,2).alignment=Alignment(horizontal='left',vertical='center')
        cov.cell(r2,3,desc).font=Font(name='Calibri',size=8,color='666666',italic=True)
        cov.cell(r2,3).alignment=Alignment(horizontal='left',vertical='center')
    for r2,txt in [(18,'  Blue = input   Black = formula   Green = cross-sheet link   Subtotals are live formulas   Y2-Y10 from model'),
                   (19,f'  Generated: {datetime.today().strftime("%d %B %Y")}  |  Trilogy Finance Suite  ·  App 02  ·  Performance Plan Builder')]:
        cov.row_dimensions[r2].height=11; cov.merge_cells(f'B{r2}:C{r2}'); c=cov[f'B{r2}']; c.value=txt
        c.font=Font(name='Calibri',size=8,color='AAAAAA',italic=True)
        c.fill=PatternFill('solid',fgColor='F5F5F5'); c.alignment=Alignment(horizontal='left',vertical='center')

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):

    def do_GET(self):
        """Health check — lets you verify the function is deployed."""
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({'status': 'ok', 'function': 'export'}).encode())

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Content-Length', '0')
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length') or 0)
            body   = self.rfile.read(length) if length > 0 else b'{}'
            data   = json.loads(body.decode('utf-8'))
            excel_bytes = build_excel(data)
            co   = str(data.get('company', 'Deal')).replace(' ', '_')
            fname = f'{co}_Performance_Plan.xlsx'
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_bytes)
        except Exception as e:
            import traceback
            err_detail = traceback.format_exc()
            payload = json.dumps({'error': str(e), 'detail': err_detail}).encode()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(payload)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(payload)

    def log_message(self, *args):
        pass
